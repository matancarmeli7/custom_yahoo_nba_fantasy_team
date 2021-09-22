import argparse

from openpyxl import load_workbook
from prettytable import PrettyTable

excel_path = "fantasy_proj_stats2021-2022.xlsx"
players_list_file_path = "team.txt"
pick_round = 0
index_table = PrettyTable(['stat', 'Elite', 'Good', 'Average', 'Below Average'])
players_table = PrettyTable(['Name', 'GP', 'Proj Rank', 'Rostered', 'FGM/A', 'FG%', 'FTM/A', 'FT%',
     '3PTM', 'PTS', 'REB', 'AST', 'ST', 'BLK', 'TO'])
total_table = PrettyTable(['Name', 'FG%', 'FT%',
     '3PTM', 'PTS', 'REB', 'AST', 'ST', 'BLK', 'TO'])
players_stats = {"FGM/A" : {"stat_number": 4, "two_values_in_one": True},
                 "FTM/A" : {"stat_number": 6, "two_values_in_one": True},
                 "3PTM"  : {"stat_number": 8},
                 "PTS"   : {"stat_number": 9},
                 "REB"   : {"stat_number": 10},
                 "AST"   : {"stat_number": 11},
                 "ST"    : {"stat_number": 12},
                 "BLK"   : {"stat_number": 13},
                 "TO"    : {"stat_number": 14}}
index_stats = [['FG%', '510', '490', '480', '465'],
               ['FT%', '835', '815', '795', '775'],
               ['3PTM', '2100', '1900', '1700', '1550'],
               ['PTS', '18000', '17000', '15500', '13500'],
               ['REB', '6150', '5550', '5300', '5100'],
               ['AST', '4250', '4100', '3950', '3700'],
               ['ST', '1100', '950', '800', '650'],
               ['BLK', '900', '750', '600', '450'],
               ['TO', '1600', '1700', '1900', '2050'],]

def index_args():
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawTextHelpFormatter,
        description='Python script to create easily customize Yahoo Fantasy NBA team')
    parser.add_argument(
        '--show-index',
        action='store_const',
        const='True',
        help='Number of times to run the given command (default: 1)',
    )
    args = parser.parse_args()
    return args

def add_player_to_table(players_table, player_rank, all_players_and_thier_stats):
    player = all_players_and_thier_stats[player_rank*2-1]
    players_table.add_row(list(player))

def get_all_players_from_excel():
    workbook = load_workbook(filename=excel_path)
    sheet = workbook.active
    all_players_and_thier_stats = sheet.iter_rows(values_only=True)
    return list(all_players_and_thier_stats)

def add_player_stats_to_total(player_rank, all_players_and_thier_stats):
    player = all_players_and_thier_stats[player_rank*2-1]
    for stat, stat_value in players_stats.items():
        if 'total1' in stat_value.keys():
            add_stat_to_total(player, stat)
        else:
            create_new_total(player, stat)

def add_stat_to_total(player, stat):
    if 'two_values_in_one' in players_stats[stat].keys():
        players_stats[stat]['total1'] += float(player[players_stats[stat]["stat_number"]].split("/")[0])
        players_stats[stat]['total2'] += float(player[players_stats[stat]["stat_number"]].split("/")[1])
    else:
        players_stats[stat]['total1'] += player[players_stats[stat]["stat_number"]]

def create_new_total(player, stat):
    if 'two_values_in_one' in players_stats[stat].keys():
        players_stats[stat]['total1'] = float(player[players_stats[stat]["stat_number"]].split("/")[0])
        players_stats[stat]['total2'] = float(player[players_stats[stat]["stat_number"]].split("/")[1])
    else:
        players_stats[stat]['total1'] = player[players_stats[stat]["stat_number"]]

def add_total_stat():
    final_total = ["Total"]
    for stat in players_stats:
        if 'two_values_in_one' in players_stats[stat].keys():
            final_total.append(int(players_stats[stat]["total1"]/players_stats[stat]["total2"]*1000))
        else:
            final_total.append(players_stats[stat]["total1"])
    total_table.add_row(final_total)

def show_indexs():
    for index_stat in index_stats:
        index_table.add_row(index_stat)
    print(index_table)
    print("\n\n")

if __name__ == "__main__":
    all_players_and_thier_stats = get_all_players_from_excel()
    players_list_file = open(players_list_file_path, "r")
    for player_number in players_list_file:
        line_strip = player_number.rstrip('\n')
        add_player_to_table(players_table, int(line_strip), all_players_and_thier_stats)
        add_player_stats_to_total(int(line_strip), all_players_and_thier_stats)
    players_list_file.close()
    args = index_args()
    if args.show_index:
        show_indexs()
    add_total_stat()
    print(total_table)
    print("\n\n")
    print(players_table)
